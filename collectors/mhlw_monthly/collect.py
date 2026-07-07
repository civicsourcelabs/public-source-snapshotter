#!/usr/bin/env python3
"""
Public-source MHLW monthly source-file collector.

This collector reads a public manifest, resolves official source files, and
optionally downloads raw ZIP/XLSX files into a source artifact.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import re
import shutil
import ssl
import sys
import tempfile
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable


OWNER_CONFIRM = "owner-approved-public-source-snapshot"
DEFAULT_USER_AGENT = (
    "Mozilla/5.0 (compatible; CivicSourceSnapshotter MHLW monthly collector; "
    "owner-approved-public-source-snapshot)"
)
ALLOWED_FETCH_TYPES = {"direct_download", "xpath", "month_context"}
ALLOWED_SOURCE_TYPES = {"todokede", "code_content", "other", "unknown"}
PRODUCT_LABEL_BY_SLUG = {"medical": "医科", "dental": "歯科", "pharmacy": "薬局"}
PRODUCT_CODE_BY_SLUG = {"medical": "01", "dental": "03", "pharmacy": "04"}
SOURCE_CATEGORY_BY_LABEL = {label: slug for slug, label in PRODUCT_LABEL_BY_SLUG.items()}
ROMANIZED_CATEGORY_PATTERNS = {
    "medical": re.compile(r"(?:^|[_\-.])ika(?:[_\-.]|heisetsu|$)", re.IGNORECASE),
    "dental": re.compile(r"(?:^|[_\-.])shika(?:[_\-.]|heisetsu|$)", re.IGNORECASE),
    "pharmacy": re.compile(r"(?:^|[_\-.])yakkyoku(?:[_\-.]|$)", re.IGNORECASE),
}
SOURCE_TITLE_BY_TYPE = {
    "todokede": "届出受理医療機関名簿",
    "code_content": "コード内容別医療機関一覧表",
}
REIWA_MONTH_RE = re.compile(r"令和([0-9０-９]+)年([0-9０-９]+)月(?:[0-9０-９]+日)?現在")
FULLWIDTH_DIGIT_TRANSLATION = str.maketrans("０１２３４５６７８９", "0123456789")


@dataclass(frozen=True)
class SourceRow:
    source_key: str
    pipeline_slug: str
    region: str
    source_label: str
    source_type: str
    fetch_type: str
    download_subdir: str
    expected_filename: str
    page_url: str = ""
    xpath: str = ""
    file_url: str = ""
    priority: int = 0


@dataclass(frozen=True)
class ResolvedSource:
    url: str
    selector: dict[str, Any]


@dataclass(frozen=True)
class SourceResult:
    row: SourceRow
    status: str
    resolved_url: str = ""
    selector: dict[str, Any] | None = None
    output_path: str = ""
    byte_size: int = 0
    sha256: str = ""
    error: str = ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--manifest", type=Path, help="MHLW monthly source manifest JSON.")
    parser.add_argument("--out-dir", type=Path, help="Output directory.")
    parser.add_argument("--source-id", default="mhlw_monthly")
    parser.add_argument("--source-snapshot-date", default="")
    parser.add_argument("--run-label", default="collector-mhlw-monthly-canary")
    parser.add_argument(
        "--artifact-mode",
        choices=("summary_only", "encrypted_full"),
        default="summary_only",
    )
    parser.add_argument("--execute", action="store_true", help="Download resolved source files.")
    parser.add_argument("--confirm", default="", help=f"Required for --execute: {OWNER_CONFIRM}")
    parser.add_argument("--pipeline-slug", default="", help="Comma-separated pipeline_slug filter.")
    parser.add_argument("--region", default="", help="Comma-separated region filter.")
    parser.add_argument("--source-type", default="", help="Comma-separated source_type filter.")
    parser.add_argument("--source-key", action="append", default=[], help="Specific source_key to include.")
    parser.add_argument("--max-sources", type=int, default=0, help="Limit selected source rows. 0 means no limit.")
    parser.add_argument("--workers", type=int, default=4, help="Reserved for workflow contract; downloads are bounded sequentially for now.")
    parser.add_argument("--pause-seconds", type=float, default=0.5)
    parser.add_argument("--timeout-seconds", type=float, default=30.0)
    parser.add_argument("--retry-count", type=int, default=2)
    parser.add_argument("--retry-backoff-seconds", type=float, default=2.0)
    parser.add_argument("--user-agent", default=DEFAULT_USER_AGENT)
    parser.add_argument(
        "--insecure-skip-tls-verify",
        action="store_true",
        help="Disable TLS verification for owner-approved troubleshooting only.",
    )
    parser.add_argument("--self-test", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.self_test:
        return run_self_test()

    if args.execute and args.confirm != OWNER_CONFIRM:
        raise SystemExit(f"--execute requires --confirm {OWNER_CONFIRM}")
    if args.manifest is None or args.out_dir is None:
        raise SystemExit("--manifest and --out-dir are required")

    manifest = load_manifest(args.manifest)
    source_snapshot_date = args.source_snapshot_date or str(manifest.get("source_snapshot_date") or "")
    source_id = args.source_id or str(manifest.get("source_id") or "mhlw_monthly")

    rows = select_rows(
        parse_source_rows(manifest),
        pipeline_slugs=csv_filter(args.pipeline_slug),
        regions=csv_filter(args.region),
        source_types=csv_filter(args.source_type),
        source_keys=set(args.source_key),
        max_sources=args.max_sources,
    )
    if not rows:
        raise SystemExit("No source rows selected")

    out_dir = args.out_dir
    prepare_output_dirs(out_dir)
    write_json(out_dir / "manifest" / "source-snapshot-manifest.json", manifest)

    results: list[SourceResult] = []
    for row in rows:
        if results:
            time.sleep(max(args.pause_seconds, 0))
        results.append(process_row(row, args=args, out_dir=out_dir))

    write_inventory(out_dir / "metrics" / "mhlw-source-file-inventory.csv", results)
    write_coverage_summary(out_dir / "metrics" / "source-coverage-summary.csv", results)
    source_units = build_source_units(
        results,
        out_dir=out_dir,
        source_id=source_id,
        source_snapshot_date=source_snapshot_date,
    )
    coverage = build_source_coverage(
        source_units,
        selected_count=len(rows),
        is_full_snapshot=is_full_snapshot(args),
        source_snapshot_date=source_snapshot_date,
    )
    write_json(out_dir / "manifest" / "mhlw-source-units.json", source_units)
    write_json(out_dir / "manifest" / "mhlw-source-coverage.json", coverage)
    write_run_manifests(
        out_dir=out_dir,
        args=args,
        source_id=source_id,
        source_snapshot_date=source_snapshot_date,
        selected_count=len(rows),
        results=results,
    )
    write_sha256sums(out_dir)

    ok = sum(1 for result in results if result.status in {"downloaded", "dry_run_resolved"})
    errors = [result for result in results if result.status == "error"]
    if errors:
        print(json.dumps({"error_details": [error_summary(result) for result in errors[:10]]}, ensure_ascii=False))
    print(
        json.dumps(
            {
                "status": "executed" if args.execute else "dry_run",
                "selected_count": len(rows),
                "ok_count": ok,
                "error_count": len(errors),
                "source_unit_ok_count": coverage["ok_units"],
                "source_unit_failed_count": coverage["failed_units"],
                "out_dir": str(out_dir),
            },
            ensure_ascii=False,
        )
    )
    if errors:
        return 1
    if coverage["is_full_snapshot"] and coverage["failed_units"]:
        print(
            json.dumps(
                {
                    "source_coverage_status": "failed",
                    "failed_units": coverage["failed_units"],
                    "missing_units": coverage["missing_units"],
                },
                ensure_ascii=False,
            )
        )
        return 1
    return 0


def error_summary(result: SourceResult) -> dict[str, str]:
    return {
        "source_key": result.row.source_key,
        "pipeline_slug": result.row.pipeline_slug,
        "region": result.row.region,
        "source_label": result.row.source_label,
        "source_type": result.row.source_type,
        "fetch_type": result.row.fetch_type,
        "expected_filename": result.row.expected_filename,
        "error": result.error,
    }


def load_manifest(path: Path) -> dict:
    manifest = json.loads(path.read_text(encoding="utf-8"))
    if manifest.get("schema_version") != "1.0":
        raise SystemExit("manifest schema_version must be 1.0")
    if not isinstance(manifest.get("source_urls"), list):
        raise SystemExit("manifest source_urls must be a list")
    return manifest


def parse_source_rows(manifest: dict) -> list[SourceRow]:
    rows: list[SourceRow] = []
    for index, raw in enumerate(manifest["source_urls"]):
        row = SourceRow(
            source_key=required(raw, "source_key"),
            pipeline_slug=required(raw, "pipeline_slug"),
            region=required(raw, "region"),
            source_label=required(raw, "source_label"),
            source_type=raw.get("source_type") or infer_source_type(raw.get("source_label", "")),
            fetch_type=required(raw, "fetch_type"),
            download_subdir=required(raw, "download_subdir"),
            expected_filename=str(raw.get("expected_filename") or ""),
            page_url=str(raw.get("page_url") or ""),
            xpath=str(raw.get("xpath") or ""),
            file_url=str(raw.get("file_url") or ""),
            priority=int(raw.get("priority") or index),
        )
        validate_row(row)
        rows.append(row)
    return sorted(rows, key=lambda row: (row.priority, row.pipeline_slug, row.region, row.source_key))


def required(raw: dict, key: str) -> str:
    value = str(raw.get(key) or "").strip()
    if not value:
        raise SystemExit(f"source row missing required key: {key}")
    return value


def validate_row(row: SourceRow) -> None:
    if row.fetch_type not in ALLOWED_FETCH_TYPES:
        raise SystemExit(f"{row.source_key}: unsupported fetch_type={row.fetch_type}")
    if row.source_type not in ALLOWED_SOURCE_TYPES:
        raise SystemExit(f"{row.source_key}: unsupported source_type={row.source_type}")
    if row.fetch_type == "direct_download" and not row.file_url:
        raise SystemExit(f"{row.source_key}: direct_download requires file_url")
    if row.fetch_type == "xpath" and not (row.page_url and row.xpath):
        raise SystemExit(f"{row.source_key}: xpath requires page_url and xpath")
    if row.fetch_type == "month_context" and not row.page_url:
        raise SystemExit(f"{row.source_key}: month_context requires page_url")
    if row.fetch_type == "month_context" and row.source_type not in {"todokede", "code_content"}:
        raise SystemExit(f"{row.source_key}: month_context requires a known monthly source_type")
    if row.fetch_type == "month_context" and row.pipeline_slug not in PRODUCT_LABEL_BY_SLUG:
        raise SystemExit(f"{row.source_key}: month_context requires a known pipeline_slug")
    safe_relative_path(row.download_subdir)
    if row.expected_filename:
        safe_relative_path(row.expected_filename)


def infer_source_type(source_label: str) -> str:
    if "届出受理" in source_label:
        return "todokede"
    if "コード内容別" in source_label:
        return "code_content"
    return "unknown"


def csv_filter(raw: str) -> set[str]:
    return {part.strip() for part in raw.split(",") if part.strip()}


def select_rows(
    rows: list[SourceRow],
    *,
    pipeline_slugs: set[str],
    regions: set[str],
    source_types: set[str],
    source_keys: set[str],
    max_sources: int,
) -> list[SourceRow]:
    selected = []
    for row in rows:
        if pipeline_slugs and row.pipeline_slug not in pipeline_slugs:
            continue
        if regions and row.region not in regions:
            continue
        if source_types and row.source_type not in source_types:
            continue
        if source_keys and row.source_key not in source_keys:
            continue
        selected.append(row)
    if max_sources > 0:
        selected = selected[:max_sources]
    return selected


def prepare_output_dirs(out_dir: Path) -> None:
    for child in ("manifest", "metrics", "checksums", "raw-files", "encrypted"):
        (out_dir / child).mkdir(parents=True, exist_ok=True)


def process_row(row: SourceRow, *, args: argparse.Namespace, out_dir: Path) -> SourceResult:
    try:
        resolved = resolve_source(row, args=args)
        if not args.execute:
            return SourceResult(
                row=row,
                status="dry_run_resolved",
                resolved_url=resolved.url,
                selector=resolved.selector,
            )

        dest = out_dir / "raw-files" / safe_relative_path(row.download_subdir) / destination_filename(row, resolved.url)
        dest.parent.mkdir(parents=True, exist_ok=True)
        tmp = dest.with_name(f"{dest.name}.tmp-{os.getpid()}")
        download_file(resolved.url, tmp, args=args)
        tmp.replace(dest)
        digest = sha256_file(dest)
        return SourceResult(
            row=row,
            status="downloaded",
            resolved_url=resolved.url,
            selector=resolved.selector,
            output_path=str(dest.relative_to(out_dir)),
            byte_size=dest.stat().st_size,
            sha256=digest,
        )
    except Exception as exc:
        return SourceResult(row=row, status="error", error=f"{type(exc).__name__}: {exc}")


def resolve_source_url(row: SourceRow, *, args: argparse.Namespace) -> str:
    return resolve_source(row, args=args).url


def resolve_source(row: SourceRow, *, args: argparse.Namespace) -> ResolvedSource:
    if row.fetch_type == "direct_download":
        return ResolvedSource(
            url=row.file_url,
            selector={
                "type": "direct_download",
                "selected_href": row.file_url,
                "selected_filename": resolved_basename(row.file_url),
            },
        )
    html_text = request_text(row.page_url, args=args)
    if row.fetch_type == "month_context":
        return resolve_month_context_href(
            html_text,
            page_url=row.page_url,
            row=row,
            source_snapshot_date=args.source_snapshot_date,
        )
    href = resolve_xpath_href(html_text, row.page_url, row.xpath)
    return ResolvedSource(
        url=href,
        selector={
            "type": "xpath",
            "page_url": row.page_url,
            "xpath": row.xpath,
            "selected_href": href,
            "selected_filename": resolved_basename(href),
        },
    )


def destination_filename(row: SourceRow, resolved_url: str) -> Path:
    if row.expected_filename:
        return safe_relative_path(row.expected_filename)
    parsed = urllib.parse.urlparse(resolved_url)
    filename = Path(urllib.parse.unquote(parsed.path)).name
    if not filename:
        filename = f"{row.source_key}.bin"
    return safe_relative_path(filename)


def resolve_xpath_href(html_text: str, page_url: str, xpath: str) -> str:
    try:
        from lxml import html as lxml_html
    except ImportError as exc:
        raise RuntimeError("xpath fetch_type requires lxml. Install lxml in the workflow.") from exc
    tree = lxml_html.fromstring(html_text)
    elements = tree.xpath(xpath)
    if not elements:
        raise RuntimeError(f"XPath did not match: {xpath}")
    first = elements[0]
    href = first.get("href") if hasattr(first, "get") else str(first).strip()
    if not href and hasattr(first, "xpath"):
        children = first.xpath(".//a[@href]")
        if children:
            href = children[0].get("href")
    if not href:
        raise RuntimeError(f"XPath matched but href was empty: {xpath}")
    return urllib.parse.urljoin(page_url, href)


def resolve_month_context_href(
    html_text: str,
    *,
    page_url: str,
    row: SourceRow,
    source_snapshot_date: str,
) -> ResolvedSource:
    try:
        from lxml import html as lxml_html
    except ImportError as exc:
        raise RuntimeError("month_context fetch_type requires lxml. Install lxml in the workflow.") from exc

    target_month = parse_target_month(source_snapshot_date)
    tree = lxml_html.fromstring(html_text)
    candidates = []
    for anchor in tree.xpath("//a[@href]"):
        text = normalize_space(anchor.text_content())
        href = urllib.parse.urljoin(page_url, anchor.get("href"))
        context_text = normalize_space(" ".join(str(part) for part in anchor.xpath("preceding::text()")))
        candidates.append(
            {
                "snapshot_month": latest_reiwa_month(context_text),
                "format": latest_source_format(context_text),
                "text": text,
                "href": href,
                "basename": resolved_basename(href),
            }
        )

    source_title = SOURCE_TITLE_BY_TYPE[row.source_type]
    product_label = PRODUCT_LABEL_BY_SLUG[row.pipeline_slug]
    matches = [
        candidate
        for candidate in candidates
        if candidate["snapshot_month"] == target_month
        and candidate["format"] == "excel"
        and source_title in candidate["text"]
        and f"（{product_label}）" in candidate["text"]
    ]
    if len(matches) != 1:
        summary = [
            {
                "snapshot_month": candidate["snapshot_month"],
                "format": candidate["format"],
                "text": candidate["text"],
                "basename": candidate["basename"],
            }
            for candidate in candidates
            if source_title in candidate["text"] and f"（{product_label}）" in candidate["text"]
        ][:10]
        raise RuntimeError(
            "month_context expected one source link; "
            f"resolved_count={len(matches)} target_month={target_month} "
            f"source_type={row.source_type} pipeline_slug={row.pipeline_slug} links={summary}"
        )
    selected = matches[0]
    validate_month_context_filename(
        row,
        basename=selected["basename"],
        target_month=target_month,
    )
    selector_candidates = [
        {
            "date": candidate["snapshot_month"],
            "format": candidate["format"],
            "text": candidate["text"],
            "href": candidate["href"],
            "filename": candidate["basename"],
        }
        for candidate in candidates
        if source_title in candidate["text"] and f"（{product_label}）" in candidate["text"]
    ][:20]
    return ResolvedSource(
        url=selected["href"],
        selector={
            "type": "month_context",
            "page_url": page_url,
            "target_month": target_month,
            "selected_text": selected["text"],
            "selected_href": selected["href"],
            "selected_filename": selected["basename"],
            "candidate_count": len(selector_candidates),
            "candidates": selector_candidates,
        },
    )


def parse_target_month(source_snapshot_date: str) -> str:
    value = normalize_digits(source_snapshot_date.strip())
    match = re.fullmatch(r"([0-9]{4})-([0-9]{2})(?:-[0-9]{2})?", value)
    if not match:
        raise RuntimeError(f"source_snapshot_date must be YYYY-MM or YYYY-MM-DD: {source_snapshot_date}")
    return f"{match.group(1)}-{match.group(2)}"


def latest_reiwa_month(text: str) -> str:
    matches = list(REIWA_MONTH_RE.finditer(normalize_digits(text)))
    if not matches:
        return ""
    last = matches[-1]
    year = 2018 + int(last.group(1))
    month = int(last.group(2))
    return f"{year:04d}-{month:02d}"


def latest_source_format(text: str) -> str:
    normalized = normalize_space(text)
    pdf_index = max(normalized.rfind("（1）PDFファイル"), normalized.rfind("(1) PDF"))
    excel_index = max(
        normalized.rfind("（2）エクセルファイル"),
        normalized.rfind("(2) Excel"),
        normalized.rfind("(2) Excel file"),
    )
    if excel_index > pdf_index:
        return "excel"
    if pdf_index >= 0:
        return "pdf"
    return ""


def normalize_digits(value: str) -> str:
    return value.translate(FULLWIDTH_DIGIT_TRANSLATION)


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def resolved_basename(url: str) -> str:
    parsed = urllib.parse.urlparse(url)
    return Path(urllib.parse.unquote(parsed.path)).name


def validate_month_context_filename(row: SourceRow, *, basename: str, target_month: str) -> None:
    code = PRODUCT_CODE_BY_SLUG[row.pipeline_slug]
    yy_month = target_month[2:4] + target_month[5:7]
    if row.source_type == "todokede":
        pattern = rf"^{re.escape(yy_month)}-06_10-{re.escape(code)}\.zip$"
    elif row.source_type == "code_content":
        pattern = rf"^{re.escape(yy_month)}-01-{re.escape(code)}(?:_[0-9]+)?\.zip$"
    else:
        raise RuntimeError(f"month_context unsupported source_type={row.source_type}")
    if not re.fullmatch(pattern, basename):
        raise RuntimeError(
            f"resolved filename did not match month_context pattern: basename={basename} pattern={pattern}"
        )


def request_text(url: str, *, args: argparse.Namespace) -> str:
    data = request_bytes(url, args=args)
    return data.decode("utf-8", errors="replace")


def download_file(url: str, dest: Path, *, args: argparse.Namespace) -> None:
    data = request_bytes(url, args=args)
    dest.write_bytes(data)


def request_bytes(url: str, *, args: argparse.Namespace) -> bytes:
    last_error: Exception | None = None
    context = ssl._create_unverified_context() if args.insecure_skip_tls_verify else None
    for attempt in range(args.retry_count + 1):
        try:
            request = urllib.request.Request(url, headers={"User-Agent": args.user_agent})
            with urllib.request.urlopen(request, timeout=args.timeout_seconds, context=context) as response:
                return response.read()
        except (urllib.error.URLError, TimeoutError, ssl.SSLError) as exc:
            last_error = exc
            if attempt >= args.retry_count:
                break
            time.sleep(args.retry_backoff_seconds * (attempt + 1))
    raise RuntimeError(f"request failed for {url}: {last_error}")


def safe_relative_path(value: str) -> Path:
    path = Path(value)
    if path.is_absolute() or any(part in {"", ".", ".."} for part in path.parts):
        raise ValueError(f"unsafe relative path: {value}")
    return path


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_inventory(path: Path, results: list[SourceResult]) -> None:
    fieldnames = [
        "source_key",
        "pipeline_slug",
        "region",
        "source_label",
        "source_type",
        "fetch_type",
        "download_subdir",
        "expected_filename",
        "status",
        "resolved_url",
        "output_path",
        "byte_size",
        "sha256",
        "error",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for result in results:
            writer.writerow(
                {
                    "source_key": result.row.source_key,
                    "pipeline_slug": result.row.pipeline_slug,
                    "region": result.row.region,
                    "source_label": result.row.source_label,
                    "source_type": result.row.source_type,
                    "fetch_type": result.row.fetch_type,
                    "download_subdir": result.row.download_subdir,
                    "expected_filename": result.row.expected_filename,
                    "status": result.status,
                    "resolved_url": result.resolved_url,
                    "output_path": result.output_path,
                    "byte_size": result.byte_size,
                    "sha256": result.sha256,
                    "error": result.error,
                }
            )


def write_coverage_summary(path: Path, results: list[SourceResult]) -> None:
    groups: dict[tuple[str, str], list[SourceResult]] = {}
    for result in results:
        groups.setdefault((result.row.pipeline_slug, result.row.source_type), []).append(result)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "pipeline_slug",
                "source_type",
                "source_count",
                "downloaded_count",
                "dry_run_resolved_count",
                "error_count",
                "total_bytes",
            ],
        )
        writer.writeheader()
        for (pipeline_slug, source_type), group in sorted(groups.items()):
            writer.writerow(
                {
                    "pipeline_slug": pipeline_slug,
                    "source_type": source_type,
                    "source_count": len(group),
                    "downloaded_count": sum(1 for r in group if r.status == "downloaded"),
                    "dry_run_resolved_count": sum(1 for r in group if r.status == "dry_run_resolved"),
                    "error_count": sum(1 for r in group if r.status == "error"),
                    "total_bytes": sum(r.byte_size for r in group),
                }
            )


def build_source_units(
    results: list[SourceResult],
    *,
    out_dir: Path,
    source_id: str,
    source_snapshot_date: str,
) -> dict[str, Any]:
    units: list[dict[str, Any]] = []
    snapshot_month = parse_target_month(source_snapshot_date)
    for result in results:
        row = result.row
        validation = {
            "filename_rule": filename_rule_status(result),
            "content_probe": "skipped",
            "observed_categories": [],
            "observed_source_kinds": [],
            "errors": [],
        }
        status = source_unit_status_from_result(result)

        if result.status == "downloaded" and result.output_path:
            probe = probe_source_file(out_dir / safe_relative_path(result.output_path))
            validation.update(probe)
            status = source_unit_status_from_validation(row, probe)

        raw_file = {
            "path": result.output_path,
            "filename": Path(result.output_path).name if result.output_path else "",
            "byte_size": result.byte_size,
            "sha256": result.sha256,
        }
        if result.status == "dry_run_resolved":
            raw_file["filename"] = destination_filename(row, result.resolved_url).as_posix()

        units.append(
            {
                "schema_version": 1,
                "source_id": source_id,
                "source_snapshot_date": source_snapshot_date,
                "snapshot_month": snapshot_month,
                "unit_id": f"{row.source_key}.{snapshot_month}",
                "source_key": row.source_key,
                "region": row.region,
                "source_label": row.source_label,
                "source_kind": row.source_type,
                "source_category": row.pipeline_slug,
                "fetch_type": row.fetch_type,
                "selector": result.selector or {},
                "raw_file": raw_file,
                "validation": validation,
                "status": status,
                "error": result.error,
            }
        )
    return {
        "schema_version": 1,
        "source_id": source_id,
        "source_snapshot_date": source_snapshot_date,
        "snapshot_month": snapshot_month,
        "unit_count": len(units),
        "units": units,
    }


def build_source_coverage(
    source_units: dict[str, Any],
    *,
    selected_count: int,
    is_full_snapshot: bool,
    source_snapshot_date: str,
) -> dict[str, Any]:
    units = source_units.get("units") if isinstance(source_units.get("units"), list) else []
    by_category_and_kind: dict[str, dict[str, dict[str, int]]] = {}
    failed = []
    missing = []
    for unit in units:
        category = str(unit.get("source_category") or "unknown")
        kind = str(unit.get("source_kind") or "unknown")
        status = str(unit.get("status") or "unknown")
        status_counts = by_category_and_kind.setdefault(category, {}).setdefault(kind, {})
        status_counts[status] = status_counts.get(status, 0) + 1
        if status == "missing":
            missing.append(unit.get("unit_id"))
        elif status not in {"ok", "dry_run_resolved"}:
            failed.append(unit.get("unit_id"))

    return {
        "schema_version": 1,
        "source_snapshot_date": source_snapshot_date,
        "snapshot_month": source_units.get("snapshot_month", ""),
        "is_full_snapshot": bool(is_full_snapshot),
        "expected_units": selected_count,
        "ok_units": sum(1 for unit in units if unit.get("status") == "ok"),
        "missing_units": len(missing),
        "failed_units": len(failed),
        "by_category_and_kind": by_category_and_kind,
        "failed_unit_ids": [str(value) for value in failed if value][:50],
        "missing_unit_ids": [str(value) for value in missing if value][:50],
    }


def is_full_snapshot(args: argparse.Namespace) -> bool:
    return (
        bool(args.execute)
        and args.artifact_mode == "encrypted_full"
        and not csv_filter(args.pipeline_slug)
        and not csv_filter(args.region)
        and not csv_filter(args.source_type)
        and not args.source_key
        and int(args.max_sources or 0) == 0
    )


def source_unit_status_from_result(result: SourceResult) -> str:
    if result.status == "downloaded":
        return "ok"
    if result.status == "dry_run_resolved":
        return "dry_run_resolved"
    return "error"


def filename_rule_status(result: SourceResult) -> str:
    if result.row.fetch_type != "month_context":
        return "not_applicable"
    if result.status == "error":
        return "failed"
    return "passed"


def probe_source_file(path: Path) -> dict[str, Any]:
    try:
        payloads = xlsx_payloads(path)
    except Exception as exc:
        return {
            "content_probe": "failed",
            "observed_categories": [],
            "observed_source_kinds": [],
            "errors": [f"{type(exc).__name__}: {exc}"],
        }
    if not payloads:
        return {
            "content_probe": "failed",
            "observed_categories": [],
            "observed_source_kinds": [],
            "errors": ["no xlsx payloads found"],
        }

    observed_categories: set[str] = set()
    observed_source_kinds: set[str] = set()
    errors: list[str] = []
    for member_name, payload in payloads:
        try:
            inspection = inspect_xlsx_payload(member_name, payload)
            observed_categories.update(inspection["categories"])
            observed_source_kinds.update(inspection["source_kinds"])
        except Exception as exc:
            errors.append(f"{member_name}: {type(exc).__name__}: {exc}")

    return {
        "content_probe": "passed" if not errors else "partial",
        "observed_categories": sorted(observed_categories),
        "observed_source_kinds": sorted(observed_source_kinds),
        "errors": errors[:5],
    }


def xlsx_payloads(path: Path) -> list[tuple[str, bytes]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return [(path.name, path.read_bytes())]
    if suffix != ".zip":
        return []
    payloads: list[tuple[str, bytes]] = []
    with zipfile.ZipFile(path) as archive:
        for member_name in archive.namelist():
            if member_name.lower().endswith(".xlsx"):
                payloads.append((member_name, archive.read(member_name)))
    return payloads


def inspect_xlsx_payload(member_name: str, payload: bytes) -> dict[str, set[str]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("content probe requires openpyxl. Install openpyxl in the workflow.") from exc

    categories: set[str] = set()
    source_kinds: set[str] = set()
    categories.update(source_categories_from_text(member_name, loose=True))
    source_kinds.update(source_kinds_from_text(member_name))
    workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            categories.update(source_categories_from_text(sheet_name, loose=True))
            source_kinds.update(source_kinds_from_text(sheet_name))
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(max_row=25, values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    value = normalize_space(str(cell).replace("\u3000", " "))
                    categories.update(source_categories_from_text(value))
                    if "現在" in value or "現存" in value or "コード内容別" in value:
                        categories.update(source_categories_from_text(value, loose=True))
                    source_kinds.update(source_kinds_from_text(value))
    finally:
        workbook.close()
    return {"categories": categories, "source_kinds": source_kinds}


def source_categories_from_text(value: str, *, loose: bool = False) -> set[str]:
    categories: set[str] = set()
    normalized = normalize_space(value).lower()
    for label, category in SOURCE_CATEGORY_BY_LABEL.items():
        if normalized == label or (loose and label in normalized):
            categories.add(category)
    for category, pattern in ROMANIZED_CATEGORY_PATTERNS.items():
        if pattern.search(normalized):
            categories.add(category)
    return categories


def source_kinds_from_text(value: str) -> set[str]:
    kinds: set[str] = set()
    if "コード内容別" in value:
        kinds.add("code_content")
    if "届出受理" in value or "受理番号" in value:
        kinds.add("todokede")
    return kinds


def source_unit_status_from_validation(row: SourceRow, probe: dict[str, Any]) -> str:
    if probe.get("content_probe") == "failed":
        return "content_probe_failed"
    categories = set(str(value) for value in probe.get("observed_categories") or [])
    if categories and row.pipeline_slug not in categories:
        return "category_mismatch"
    source_kinds = set(str(value) for value in probe.get("observed_source_kinds") or [])
    if source_kinds and row.source_type not in source_kinds:
        return "source_kind_mismatch"
    return "ok"


def write_run_manifests(
    *,
    out_dir: Path,
    args: argparse.Namespace,
    source_id: str,
    source_snapshot_date: str,
    selected_count: int,
    results: list[SourceResult],
) -> None:
    downloaded_count = sum(1 for result in results if result.status == "downloaded")
    dry_run_count = sum(1 for result in results if result.status == "dry_run_resolved")
    error_count = sum(1 for result in results if result.status == "error")
    metrics = {
        "schema_version": "1.0",
        "source_id": source_id,
        "source_snapshot_date": source_snapshot_date,
        "run_label": args.run_label,
        "artifact_mode": args.artifact_mode,
        "execute": bool(args.execute),
        "selected_count": selected_count,
        "downloaded_count": downloaded_count,
        "dry_run_resolved_count": dry_run_count,
        "error_count": error_count,
        "total_bytes": sum(result.byte_size for result in results),
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }
    run_manifest = {
        "schema_version": "1.0",
        "collector_repo": os.environ.get("GITHUB_REPOSITORY", ""),
        "source_id": source_id,
        "source_snapshot_date": source_snapshot_date,
        "run_label": args.run_label,
        "github_run_id": os.environ.get("GITHUB_RUN_ID", ""),
        "github_run_url": github_run_url(),
        "completed_at": metrics["generated_at"],
        "artifact_mode": args.artifact_mode,
        "selected_count": selected_count,
        "execute": bool(args.execute),
    }
    write_json(out_dir / "metrics" / "fetch-metrics.json", metrics)
    write_json(out_dir / "manifest" / "collector-run-manifest.json", run_manifest)


def github_run_url() -> str:
    server = os.environ.get("GITHUB_SERVER_URL")
    repo = os.environ.get("GITHUB_REPOSITORY")
    run_id = os.environ.get("GITHUB_RUN_ID")
    if not (server and repo and run_id):
        return ""
    return f"{server}/{repo}/actions/runs/{run_id}"


def write_sha256sums(out_dir: Path) -> None:
    lines = []
    for path in sorted(out_dir.rglob("*")):
        if not path.is_file() or path.relative_to(out_dir).as_posix() == "checksums/SHA256SUMS":
            continue
        lines.append(f"{sha256_file(path)}  {path.relative_to(out_dir).as_posix()}")
    (out_dir / "checksums" / "SHA256SUMS").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def run_self_test() -> int:
    with tempfile.TemporaryDirectory() as tmp_dir_raw:
        tmp_dir = Path(tmp_dir_raw)
        payload = tmp_dir / "source.zip"
        payload.write_bytes(b"self-test-source")
        manifest = {
            "schema_version": "1.0",
            "source_id": "mhlw_monthly",
            "source_name": "self-test",
            "source_snapshot_date": "2026-04-01",
            "source_urls": [
                {
                    "source_key": "self-test-direct",
                    "pipeline_slug": "medical",
                    "region": "self",
                    "source_label": "届出受理_self",
                    "source_type": "todokede",
                    "fetch_type": "direct_download",
                    "file_url": payload.as_uri(),
                    "download_subdir": "self/届出受理",
                    "expected_filename": "source.zip",
                }
            ],
        }
        manifest_path = tmp_dir / "manifest.json"
        out_dir = tmp_dir / "out"
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False), encoding="utf-8")
        args = argparse.Namespace(
            manifest=manifest_path,
            out_dir=out_dir,
            source_id="mhlw_monthly",
            source_snapshot_date="2026-04-01",
            run_label="self-test",
            artifact_mode="summary_only",
            execute=True,
            confirm=OWNER_CONFIRM,
            pipeline_slug="",
            region="",
            source_type="",
            source_key=[],
            max_sources=0,
            workers=1,
            pause_seconds=0,
            timeout_seconds=5,
            retry_count=0,
            retry_backoff_seconds=0,
            user_agent=DEFAULT_USER_AGENT,
            insecure_skip_tls_verify=False,
            self_test=False,
        )
        manifest_loaded = load_manifest(args.manifest)
        rows = parse_source_rows(manifest_loaded)
        prepare_output_dirs(out_dir)
        results = [process_row(rows[0], args=args, out_dir=out_dir)]
        write_inventory(out_dir / "metrics" / "mhlw-source-file-inventory.csv", results)
        write_coverage_summary(out_dir / "metrics" / "source-coverage-summary.csv", results)
        write_run_manifests(
            out_dir=out_dir,
            args=args,
            source_id=args.source_id,
            source_snapshot_date=args.source_snapshot_date,
            selected_count=1,
            results=results,
        )
        write_sha256sums(out_dir)
        copied = out_dir / "raw-files" / "self" / "届出受理" / "source.zip"
        assert copied.read_bytes() == b"self-test-source"
        assert results[0].status == "downloaded"
        assert (out_dir / "checksums" / "SHA256SUMS").exists()
        month_page = tmp_dir / "month-page.html"
        month_page.write_text(
            """
            <html><body>
              <p>（2）エクセルファイル</p>
              <p>令和8年7月1日現在</p>
              <ul>
                <li><a href="2607-06_10-01.zip">届出受理医療機関名簿（医科）［ZIP形式］</a></li>
                <li><a href="2607-06_10-03.zip">届出受理医療機関名簿（歯科）［ZIP形式］</a></li>
              </ul>
              <p>令和8年6月1日現在</p>
              <ul>
                <li><a href="2606-06_10-01.zip">届出受理医療機関名簿（医科）［ZIP形式］</a></li>
              </ul>
            </body></html>
            """,
            encoding="utf-8",
        )
        month_args = argparse.Namespace(**{**vars(args), "source_snapshot_date": "2026-07-01"})
        month_row = SourceRow(
            source_key="self-test-month",
            pipeline_slug="medical",
            region="self",
            source_label="届出受理_self",
            source_type="todokede",
            fetch_type="month_context",
            download_subdir="self/届出受理",
            expected_filename="",
            page_url=month_page.as_uri(),
        )
        assert resolve_source_url(month_row, args=month_args).endswith("/2607-06_10-01.zip")
        code_page = tmp_dir / "code-page.html"
        code_page.write_text(
            """
            <html><body>
              <p>令和8年7月1日現在</p>
              <p>（2）エクセルファイル</p>
              <ul>
                <li><a href="2607-01-01_2.zip">コード内容別医療機関一覧表（医科）［ZIP形式］</a></li>
              </ul>
            </body></html>
            """,
            encoding="utf-8",
        )
        code_row = SourceRow(
            source_key="self-test-code",
            pipeline_slug="medical",
            region="self",
            source_label="コード内容別_self",
            source_type="code_content",
            fetch_type="month_context",
            download_subdir="self/コード内容別",
            expected_filename="",
            page_url=code_page.as_uri(),
        )
        assert resolve_source_url(code_row, args=month_args).endswith("/2607-01-01_2.zip")
        print(json.dumps({"status": "self_test_ok", "out_dir": str(out_dir)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
